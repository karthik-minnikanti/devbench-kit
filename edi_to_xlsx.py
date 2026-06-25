#!/usr/bin/env python3
"""
EDI to XLSX Converter
Reads EDI file, parses each line by * delimiter, and creates an XLSX file
with segment names and all field values, preserving empty fields.
"""

import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def parse_edi_line(line):
    """Parse an EDI line by * delimiter, preserving empty fields."""
    # Remove trailing whitespace and newline
    line = line.rstrip('\r\n').rstrip()
    
    if not line:
        return None
    
    # Split by * delimiter
    fields = line.split('*')
    
    # Trim whitespace from each field (whitespace-only fields become empty)
    fields = [field.strip() if field.strip() else '' for field in fields]
    
    return fields

def create_xlsx(edi_file, output_file):
    """Read EDI file and create XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EDI Data"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Find maximum number of fields across all lines to set column headers
    max_fields = 0
    all_data = []
    
    # First pass: read all data and find max fields
    with open(edi_file, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            fields = parse_edi_line(line)
            if fields:
                all_data.append(fields)
                max_fields = max(max_fields, len(fields))
    
    if max_fields == 0:
        print("Error: No data found in EDI file!")
        return False
    
    # Create headers
    headers = ['Segment'] + [f'Field{i}' for i in range(1, max_fields)]
    ws.append(headers)
    
    # Apply header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Write data rows
    for fields in all_data:
        # Segment name is first field, rest are field values
        row = [fields[0]]  # Segment name
        # Add remaining fields (Field[1] onwards)
        for i in range(1, max_fields):
            if i < len(fields):
                row.append(fields[i])
            else:
                row.append('')  # Empty cell if field doesn't exist
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"Successfully created XLSX file: {output_file}")
    print(f"Total rows: {len(all_data)}")
    print(f"Total columns: {max_fields}")
    return True

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 edi_to_xlsx.py <edi_file> [output_file]")
        print("Example: python3 edi_to_xlsx.py /Users/karthikminnikanti/Downloads/X291-ambulance.edi")
        sys.exit(1)
    
    edi_file = sys.argv[1]
    
    # Generate output filename if not provided
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # Use same name as input file but with .xlsx extension
        if edi_file.endswith('.edi'):
            output_file = edi_file[:-4] + '.xlsx'
        else:
            output_file = edi_file + '.xlsx'
    
    try:
        create_xlsx(edi_file, output_file)
    except FileNotFoundError:
        print(f"Error: File '{edi_file}' not found!")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    main()


